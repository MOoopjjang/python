#!python3
# -*- coding:utf-8 -*-


from openpyxl import load_workbook
from email.mime.text import MIMEText
from email.mime.multipart import  MIMEMultipart
from smtplib import SMTP_SSL
from email.mime.base import MIMEBase
from email import encoders
from os.path import basename
import os

SMTP_SERVER = 'smtp.naver.com'
SMTP_PORT = 465
SMTP_USER = 'xferlog@naver.com'
SMTP_PASSWORD = "wpswkd1gkf"



def send_mail(name , addr = '' , cc = '' , hcc = '' , contents = '' , attachment = False):
    msg = MIMEMultipart('alternative')

    if attachment:
        msg = MIMEMultipart('mixed')

    text = MIMEText(contents)
    print(f'text : {text}')

    msg['From'] = SMTP_USER
    msg['To'] = addr
    msg['CC'] = cc
    msg['Subject'] = name+'님 , 메일이 도착했습니다.'
    msg.attach(text)

    if attachment:
        file_data = MIMEBase('application' , 'octet-stream')
        with open(attachment , 'rb') as r:
            file_contents = r.read()
            file_data.set_payload(file_contents)
            encoders.encode_base64(file_data)

            filename = basename(attachment)
            file_data.add_header('Content-Disposition' , 'attachment' , filename = filename)
            msg.attach(file_data)

    taddr = ','.join((addr, cc, hcc))
    smtp = SMTP_SSL(SMTP_SERVER , SMTP_PORT)
    smtp.login(SMTP_USER , SMTP_PASSWORD)
    smtp.sendmail('xferlog@naver.com' , taddr.split(',') , msg.as_string())
    smtp.close()


if __name__ == '__main__':
    xlsx = load_workbook('수강생_결제정보.xlsx', read_only=True)
    sheet = xlsx.active

    for row in sheet.iter_rows():
        name = row[0].value
        mail = row[1].value
        status = row[3].value

        if status == '결제완료':
            contents = '결제 완료가 확인되어 커리큘럼을 안내해 드립니다.'
            send_mail(name, addr =  mail, contents = contents, attachment = os.getcwd()+'/schedule.xlsx')
